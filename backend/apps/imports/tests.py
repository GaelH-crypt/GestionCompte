import io
import datetime
import openpyxl
from django.test import TestCase
from apps.imports.services.parser import parse_excel, ColumnMappingRequired, _to_float
from apps.imports.services.categorizer import suggest_category


class ToFloatTest(TestCase):
    def test_plain_float(self):
        self.assertAlmostEqual(_to_float(1234.56), 1234.56)

    def test_comma_decimal(self):
        self.assertAlmostEqual(_to_float('1234,56'), 1234.56)

    def test_comma_thousands_dot_decimal(self):
        self.assertAlmostEqual(_to_float('1,234.56'), 1234.56)

    def test_dot_thousands_comma_decimal(self):
        self.assertAlmostEqual(_to_float('1.234,56'), 1234.56)

    def test_nbsp_thousands(self):
        self.assertAlmostEqual(_to_float('1\xa0234,56'), 1234.56)

    def test_none_returns_none(self):
        self.assertIsNone(_to_float(None))

    def test_nan_returns_none(self):
        import math
        self.assertIsNone(_to_float(float('nan')))

    def test_text_returns_none(self):
        self.assertIsNone(_to_float('abc'))


def _make_excel():
    wb = openpyxl.Workbook()

    # Feuille "Vos comptes"
    ws = wb.active
    ws.title = 'Vos comptes'
    ws.append([None])
    ws.append(['Compte', 'R.I.B.', 'Solde', 'Dev'])
    ws.append(['C/C EUROCOMPTE CONFORT', '10278 02625 00022060507', 156.22, 'EUR'])
    ws.append(['LIVRET BLEU', '10278 02625 00023120602', 10.0, 'EUR'])

    # Feuille compte courant
    wc = wb.create_sheet('Cpt 02625 00022060507')
    wc.append(['R.I.B. : 10278 02625 00022060507'] + [None] * 6)
    wc.append([None, None, None, 'Solde initial :', 'Solde initial :', None, 'EUR'])
    wc.append(['Liste de vos comptes'] * 6 + [None])
    wc.append(['Date', 'Valeur', 'Libellé', 'Débit', 'Crédit', 'Solde', 'Dev'])
    wc.append([datetime.datetime(2026, 5, 1), datetime.datetime(2026, 5, 1), 'CARREFOUR MARKET', -42.5, None, None, 'EUR'])
    wc.append([datetime.datetime(2026, 5, 2), datetime.datetime(2026, 5, 2), 'SALAIRE MAI', None, 2500.0, None, 'EUR'])
    wc.append([None, None, None, 'Solde au 28/05/2026 :', 'Solde au 28/05/2026 :', 156.22, 'EUR'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ParserTest(TestCase):
    def setUp(self):
        self.file = _make_excel()

    def test_accounts_extracted(self):
        result = parse_excel(self.file)
        self.assertEqual(len(result['accounts']), 2)
        acc = result['accounts'][0]
        self.assertEqual(acc['name'], 'C/C EUROCOMPTE CONFORT')
        self.assertEqual(acc['rib'], '10278 02625 00022060507')
        self.assertAlmostEqual(float(acc['balance']), 156.22)

    def test_transactions_extracted(self):
        result = parse_excel(self.file)
        txs = result['transactions']['10278 02625 00022060507']
        self.assertEqual(len(txs), 2)

    def test_debit_transaction(self):
        result = parse_excel(self.file)
        txs = result['transactions']['10278 02625 00022060507']
        debit = next(t for t in txs if t['transaction_type'] == 'expense')
        self.assertEqual(debit['description'], 'CARREFOUR MARKET')
        self.assertAlmostEqual(float(debit['amount']), 42.5)
        self.assertEqual(debit['date'], '2026-05-01')

    def test_credit_transaction(self):
        result = parse_excel(self.file)
        txs = result['transactions']['10278 02625 00022060507']
        credit = next(t for t in txs if t['transaction_type'] == 'income')
        self.assertAlmostEqual(float(credit['amount']), 2500.0)

    def test_footer_rows_skipped(self):
        result = parse_excel(self.file)
        txs = result['transactions']['10278 02625 00022060507']
        # La ligne "Solde au ..." ne doit pas apparaître
        self.assertFalse(any('Solde au' in t['description'] for t in txs))

    def test_rib_consistency(self):
        result = parse_excel(self.file)
        account_ribs = {a['rib'] for a in result['accounts']}
        transaction_ribs = set(result['transactions'].keys())
        # All transaction RIBs must correspond to an account RIB
        self.assertTrue(transaction_ribs.issubset(account_ribs))

    def test_parse_excel_accepts_bytes(self):
        result = parse_excel(_make_excel().read())
        self.assertIn('accounts', result)
        self.assertEqual(len(result['accounts']), 2)


from django.contrib.auth.models import User
from apps.accounts.models import Account
from apps.transactions.models import Transaction as TxModel
from apps.imports.services.deduplicator import filter_duplicates


class DeduplicatorTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('dup_user', password='pass')
        self.account = Account.objects.create(
            user=self.user, name='Test', account_type='checking',
            initial_balance=0, color='#fff', icon='CreditCard'
        )
        TxModel.objects.create(
            user=self.user, account=self.account,
            transaction_type='expense', amount='42.50',
            description='CARREFOUR MARKET', date='2026-05-01'
        )

    def test_existing_transaction_detected(self):
        candidates = [
            {'date': '2026-05-01', 'description': 'CARREFOUR MARKET', 'amount': 42.50, 'transaction_type': 'expense'},
            {'date': '2026-05-02', 'description': 'SALAIRE MAI', 'amount': 2500.0, 'transaction_type': 'income'},
        ]
        new_txs, dup_count = filter_duplicates(candidates, self.account.id)
        self.assertEqual(dup_count, 1)
        self.assertEqual(len(new_txs), 1)
        self.assertEqual(new_txs[0]['description'], 'SALAIRE MAI')

    def test_no_duplicates(self):
        candidates = [
            {'date': '2026-06-01', 'description': 'LIDL', 'amount': 15.0, 'transaction_type': 'expense'},
        ]
        new_txs, dup_count = filter_duplicates(candidates, self.account.id)
        self.assertEqual(dup_count, 0)
        self.assertEqual(len(new_txs), 1)


class CategorizerTest(TestCase):
    def test_supermarket(self):
        self.assertEqual(suggest_category('PAIEMENT CARREFOUR MARKET CARTE'), 'Alimentation')

    def test_rent(self):
        self.assertEqual(suggest_category('PRLV SEPA OPH DE CALAIS CREANCES LOCATIVES'), 'Logement')

    def test_energy(self):
        self.assertEqual(suggest_category('PRELEVEMENT EDF PARTICULIERS'), 'Factures')

    def test_transport(self):
        self.assertEqual(suggest_category('PAIEMENT SNCF BILLETS'), 'Transport')

    def test_salary(self):
        self.assertEqual(suggest_category('VIREMENT SALAIRE MARS'), 'Revenus')

    def test_no_match(self):
        self.assertIsNone(suggest_category('OPERATION DIVERSE'))


def _make_generic_single_amount_excel() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Compte Courant'
    ws.append(['Date', 'Libellé', 'Montant'])
    ws.append([datetime.datetime(2026, 5, 1), 'LIDL SUPERMARCHE', -25.50])
    ws.append([datetime.datetime(2026, 5, 2), 'VIREMENT SALAIRE', 2000.0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_generic_debit_credit_excel() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mouvements'
    ws.append(['Date', 'Description', 'Débit', 'Crédit'])
    ws.append([datetime.datetime(2026, 5, 1), 'LOYER MAI', 800.0, None])
    ws.append([datetime.datetime(2026, 5, 3), 'SALAIRE', None, 2500.0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_unknown_format_excel() -> bytes:
    """Fichier sans aucune colonne reconnue."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['Foo', 'Bar', 'Baz'])
    ws.append(['x', 'y', 'z'])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class GenericParserTest(TestCase):
    def test_single_amount_column(self):
        result = parse_excel(_make_generic_single_amount_excel())
        self.assertIn('Compte Courant', result['transactions'])
        txs = result['transactions']['Compte Courant']
        self.assertEqual(len(txs), 2)
        expense = next(t for t in txs if t['transaction_type'] == 'expense')
        self.assertAlmostEqual(expense['amount'], 25.50)
        income = next(t for t in txs if t['transaction_type'] == 'income')
        self.assertAlmostEqual(income['amount'], 2000.0)

    def test_debit_credit_columns(self):
        result = parse_excel(_make_generic_debit_credit_excel())
        self.assertIn('Mouvements', result['transactions'])
        txs = result['transactions']['Mouvements']
        self.assertEqual(len(txs), 2)
        loyer = next(t for t in txs if 'LOYER' in t['description'])
        self.assertEqual(loyer['transaction_type'], 'expense')
        salaire = next(t for t in txs if 'SALAIRE' in t['description'])
        self.assertEqual(salaire['transaction_type'], 'income')

    def test_unknown_format_raises_column_mapping_required(self):
        with self.assertRaises(ColumnMappingRequired) as ctx:
            parse_excel(_make_unknown_format_excel())
        self.assertGreater(len(ctx.exception.sheets), 0)
        sheet = ctx.exception.sheets[0]
        self.assertIn('name', sheet)
        self.assertIn('columns', sheet)
        self.assertIn('sample_rows', sheet)

    def test_column_hints_parse_transactions(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data'
        ws.append(['Timestamp', 'Note', 'Valeur'])
        ws.append([datetime.datetime(2026, 5, 1), 'Paiement CB', -50.0])
        ws.append([datetime.datetime(2026, 5, 2), 'Virement', 1000.0])
        buf = io.BytesIO()
        wb.save(buf)
        hints = {'sheet_name': 'Data', 'date_col': 0, 'description_col': 1, 'amount_col': 2}
        result = parse_excel(buf.getvalue(), column_hints=hints)
        self.assertIn('Data', result['transactions'])
        self.assertEqual(len(result['transactions']['Data']), 2)

    def test_credit_mutuel_still_works(self):
        result = parse_excel(_make_excel())
        self.assertEqual(len(result['accounts']), 2)


def _make_minimal_excel() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vos comptes'
    ws.append([None])
    ws.append(['Compte', 'R.I.B.', 'Solde', 'Dev'])
    ws.append(['C/C TEST', '10278 00000 00000000001', 100.0, 'EUR'])

    wc = wb.create_sheet('Cpt 00000 00000000001')
    wc.append(['R.I.B. : 10278 00000 00000000001'] + [None] * 6)
    wc.append([None] * 7)
    wc.append(['Liste de vos comptes'] * 6 + [None])
    wc.append(['Date', 'Valeur', 'Libellé', 'Débit', 'Crédit', 'Solde', 'Dev'])
    wc.append([datetime.datetime(2026, 4, 1), datetime.datetime(2026, 4, 1), 'LIDL', -15.0, None, None, 'EUR'])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


from rest_framework.test import APIClient
from rest_framework import status


class PreviewAPITest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user('prev_user', password='pass')
        self.client.force_authenticate(user=self.user)

    def test_preview_returns_accounts_and_transactions(self):
        excel_bytes = _make_minimal_excel()
        resp = self.client.post(
            '/api/import/preview/',
            {'file': io.BytesIO(excel_bytes)},
            format='multipart'
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertIn('accounts', data)
        self.assertIn('transactions', data)
        self.assertEqual(len(data['accounts']), 1)
        self.assertEqual(data['accounts'][0]['name'], 'C/C TEST')

    def test_preview_suggests_category(self):
        excel_bytes = _make_minimal_excel()
        resp = self.client.post(
            '/api/import/preview/',
            {'file': io.BytesIO(excel_bytes)},
            format='multipart'
        )
        txs = resp.json()['transactions']['10278 00000 00000000001']
        self.assertEqual(txs[0]['suggested_category'], 'Alimentation')

    def test_preview_requires_auth(self):
        client = APIClient()
        resp = client.post('/api/import/preview/', {}, format='multipart')
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_preview_rejects_non_xlsx(self):
        resp = self.client.post(
            '/api/import/preview/',
            {'file': io.BytesIO(b'not excel')},
            format='multipart'
        )
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)


import json


class PreviewColumnMappingAPITest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user('colmap_user', password='pass')
        self.client.force_authenticate(user=self.user)

    def test_unknown_format_returns_422(self):
        resp = self.client.post(
            '/api/import/preview/',
            {'file': io.BytesIO(_make_unknown_format_excel())},
            format='multipart',
        )
        self.assertEqual(resp.status_code, 422)
        data = resp.json()
        self.assertEqual(data['error'], 'column_mapping_required')
        self.assertIn('sheets', data)
        self.assertGreater(len(data['sheets']), 0)
        sheet = data['sheets'][0]
        self.assertIn('name', sheet)
        self.assertIn('columns', sheet)
        self.assertIn('sample_rows', sheet)

    def test_generic_format_returns_200(self):
        resp = self.client.post(
            '/api/import/preview/',
            {'file': io.BytesIO(_make_generic_single_amount_excel())},
            format='multipart',
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertIn('accounts', data)
        self.assertIn('transactions', data)
        self.assertEqual(len(data['accounts']), 1)

    def test_column_hints_accepted(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data'
        ws.append(['Timestamp', 'Note', 'Valeur'])
        ws.append([datetime.datetime(2026, 5, 1), 'Paiement CB', -50.0])
        ws.append([datetime.datetime(2026, 5, 2), 'Virement', 1000.0])
        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        hints = json.dumps({'sheet_name': 'Data', 'date_col': 0, 'description_col': 1, 'amount_col': 2})
        resp = self.client.post(
            '/api/import/preview/',
            {'file': io.BytesIO(excel_bytes), 'column_hints': hints},
            format='multipart',
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(len(data['accounts']), 1)
        self.assertEqual(len(data['transactions']['Data']), 2)
